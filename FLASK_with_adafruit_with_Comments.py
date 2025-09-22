# --- Libraries Used ---
import cv2                  # For computer vision tasks (YOLO object detection, image decoding)
import numpy as np          # For numerical operations, image arrays
import serial               # For serial communication with ESP32-CAM
import time                 # For delays, timestamps
from collections import defaultdict   # For counting detected objects easily
from flask import Flask, render_template, Response, jsonify  # For web dashboard
import threading            # For running Flask/web server in parallel
from datetime import datetime   # For adding timestamps to detections
import pandas as pd         # For Excel logging
import os                   # For checking if file exists
from openpyxl import load_workbook   # For formatting Excel sheets
from openpyxl.styles import Alignment   # For wrapping text in Excel cells
from supabase import create_client, Client   # For storing detection logs in Supabase cloud
import paho.mqtt.client as mqtt       # For publishing detection results via MQTT
import ssl                   # For secure MQTT connection (TLS)

# --- Supabase Configuration ---
SUPABASE_URL = "https://osrduqrhujdyvdbciccw.supabase.co/"   # Supabase project URL

# Secret API key for Supabase
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9zcmR1cXJodWpkeXZkYmNpY2N3Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc1Njk2NDMyOSwiZXhwIjoyMDcyNTQwMzI5fQ.B3mBxMVL0LW7pCMYdXLZAyau0sR4QpmtuUsjKB5mBxY"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)   # Create Supabase client object

# --- Serial Configuration ---
ser = serial.Serial('COM9', 115200, timeout=5)  # Open serial port for ESP32 (Windows COM9)
time.sleep(1)                                   # Give time to stabilize connection

'''
Alternative serial config for Linux/WSL
serial_port = "/dev/ttyUSB0"
ser = serial.Serial(serial_port, 115200, timeout=5)
time.sleep(1)
'''

# --- YOLO Parameters ---
whT = 224             # Width/height of YOLO input image
confThreshold = 0.3   # Confidence threshold (ignore weak detections)
nmsThreshold = 0.3    # Non-max suppression threshold (avoid duplicate boxes)

# Load class labels (coco dataset object names)
with open("coco.names", 'rt') as f:
    classNames = f.read().rstrip('\n').split('\n')

# Load YOLO model (config + pre-trained weights)
net = cv2.dnn.readNetFromDarknet("yolov3.cfg", "yolov3.weights")
net.setPreferableBackend(cv2.dnn.DNN_BACKEND_OPENCV)   # Use OpenCV backend
net.setPreferableTarget(cv2.dnn.DNN_TARGET_CPU)        # Run on CPU

# --- Flask App Setup ---
app = Flask(__name__)     # Create Flask web app
latest_result = {         # Dictionary to store latest detection results
    "detected_html": "Waiting...",  # Detected objects (for dashboard)
    "missing_html": "",             # Missing objects (for dashboard)
    "prev_img": None,               # Previous image
    "curr_img": None,               # Current image
    "prev_time": None,              # Previous timestamp
    "curr_time": None               # Current timestamp
}
lock = threading.Lock()   # Lock for safely updating latest_result across threads

# --- Excel Logging ---
EXCEL_FILE = "detection_log.xlsx"   # Excel file to save detections

# Helper: join missing objects into string
def _safe_join_missing(missing_list):
    return ", ".join(missing_list) if missing_list else "No missing objects"

# Helper: join detected objects into string
def _safe_join_detected(counts_dict):
    return ", ".join([f"{k}:{v}" for k, v in counts_dict.items()]) if counts_dict else "Nothing detected"

# Function to log results into Excel + Supabase
def log_to_excel_and_supabase(prev_time, curr_time, missing_objs, current_counts):
    # Extract date and times
    date_str = curr_time.split(" ")[0] if curr_time else ""
    prev_time_no_date = prev_time.split(" ")[1] if prev_time else ""
    curr_time_no_date = curr_time.split(" ")[1] if curr_time else ""

    # Format strings
    missing_str = _safe_join_missing(missing_objs)
    detected_str = _safe_join_detected(current_counts)
    summary = (
        f"At {curr_time_no_date} Missing: {missing_str}\n"
        f"Newly Detected: {detected_str}"
    )

    # Create new row of data
    row = {
        "Date": date_str,
        "Previous Image Timestamp": prev_time_no_date,
        "Current Image Timestamp": curr_time_no_date,
        "Missing Objects": missing_str,
        "Newly Detected": detected_str,
        "Summary": summary
    }

    # Append row into Excel
    if os.path.exists(EXCEL_FILE):
        try:
            df = pd.read_excel(EXCEL_FILE)
        except Exception:
            df = pd.DataFrame()
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    else:
        df = pd.DataFrame([row])
    df.to_excel(EXCEL_FILE, index=False)

    # Format Excel (wrap text in summary column)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for cell in ws["F"]:
        cell.alignment = Alignment(wrapText=True)
    wb.save(EXCEL_FILE)

    # Insert row into Supabase cloud database
    try:
        supabase.table("detection_log_excel").insert({
            "date": date_str,
            "previous_image_timestamp": prev_time_no_date,
            "current_image_timestamp": curr_time_no_date,
            "missing_objects": missing_str,
            "newly_detected": detected_str,
            "summary": summary
        }).execute()
        print("[Supabase] Row inserted successfully into detection_log_excel")
    except Exception as e:
        print("[Supabase] Insert error:", e)

# --- Secure MQTT Integration (Adafruit IO) ---
BROKER = "io.adafruit.com"   # Adafruit IO MQTT broker
PORT = 8883                  # Secure TLS port
USERNAME = "nishanth_11"     # Your Adafruit username
PASSWORD = "..."             # Your Adafruit AIO key
FEED = "yolo-detection"      # Feed name in Adafruit IO
TOPIC = f"{USERNAME}/feeds/{FEED}"  # Full MQTT topic path

# Create MQTT client and connect securely
mqtt_client = mqtt.Client(client_id="pc-publisher", protocol=mqtt.MQTTv311)
mqtt_client.username_pw_set(USERNAME, PASSWORD)
mqtt_client.tls_set()               # Enable TLS encryption
mqtt_client.tls_insecure_set(False) # Verify server certificate
mqtt_client.connect(BROKER, PORT, keepalive=60)
mqtt_client.loop_start()            # Start MQTT loop in background
print(f"[MQTT] Securely connected to Adafruit IO ({BROKER}), publishing on {TOPIC}")

# Helper to publish message to MQTT
def publish_mqtt(summary):
    try:
        mqtt_client.publish(TOPIC, summary)
        print(f"[MQTT] Published to Adafruit IO: {summary}")
    except Exception as e:
        print("[MQTT Error]", e)

# --- Helpers for Dashboard ---
def format_for_dashboard(counts_dict):
    return "<br>".join([f"{k}:{v}" for k, v in counts_dict.items()]) if counts_dict else "None"

def format_missing_for_dashboard(missing_list):
    return "<br>".join(missing_list) if missing_list else "None"

# --- Flask Routes ---
@app.route('/')
def index():
    return render_template("index4.html")  # Show main dashboard page

@app.route('/data')
def data():
    # Return JSON of latest detection data
    with lock:
        result = {
            "detected_html": latest_result["detected_html"],
            "missing_html": latest_result["missing_html"],
            "prev_img": "/prev.jpg" if latest_result["prev_img"] is not None else "",
            "curr_img": "/curr.jpg" if latest_result["curr_img"] is not None else "",
            "prev_time": latest_result["prev_time"] if latest_result["prev_time"] else "",
            "curr_time": latest_result["curr_time"] if latest_result["curr_time"] else ""
        }
    return jsonify(result)

@app.route('/prev.jpg')
def prev_img():
    # Serve previous image to dashboard
    with lock:
        if latest_result["prev_img"] is None:
            return "No previous image", 404
        _, buffer = cv2.imencode('.jpg', latest_result["prev_img"])
    return Response(buffer.tobytes(), mimetype='image/jpeg')

@app.route('/curr.jpg')
def curr_img():
    # Serve current image to dashboard
    with lock:
        if latest_result["curr_img"] is None:
            return "No current image", 404
        _, buffer = cv2.imencode('.jpg', latest_result["curr_img"])
    return Response(buffer.tobytes(), mimetype='image/jpeg')

# --- Serial Communication Robust Handling ---
def request_image():
    # Ask ESP32 to capture image
    try:
        ser.reset_input_buffer()
        ser.write(b"CAPTURE\n")
    except serial.SerialException as e:
        print("[Serial Error during request_image]:", e)
        reset_serial()

def receive_image():
    # Receive image data from ESP32
    try:
        # First 4 bytes = length of image
        length_bytes = ser.read(4)
        if len(length_bytes) != 4:
            print("[Serial] Failed to read image length")
            return None

        img_len = int.from_bytes(length_bytes, 'little')   # Convert to int
        img_data = bytearray()
        start_time = time.time()
        timeout = 10   # Max wait = 10 seconds

        # Read image data until complete
        while len(img_data) < img_len:
            if time.time() - start_time > timeout:
                print("[Serial] Image reception timed out")
                return None
            packet = ser.read(img_len - len(img_data))
            if not packet:
                time.sleep(0.1)
                continue
            img_data.extend(packet)

        if len(img_data) != img_len:
            print("[Serial] Incomplete image received")
            return None

        # Convert raw bytes → NumPy array → OpenCV image
        img_np = np.frombuffer(img_data, dtype=np.uint8)
        return cv2.imdecode(img_np, cv2.IMREAD_COLOR)

    except serial.SerialException as e:
        print("[Serial Exception]:", e)
        reset_serial()
        return None

def reset_serial():
    # Reset serial port if something goes wrong
    global ser
    try:
        ser.close()
    except Exception:
        pass
    time.sleep(1)
    try:
        ser = serial.Serial('COM9', 115200, timeout=5)
        time.sleep(1)
        print("[Serial] Port reset successful")
    except Exception as e:
        print("[Serial] Failed to reset serial port:", e)

# --- YOLO Detection ---
def detect_and_count(img):
    # Preprocess image for YOLO
    blob = cv2.dnn.blobFromImage(img, 1/255, (whT, whT), [0, 0, 0], 1, crop=False)
    net.setInput(blob)
    layernames = net.getLayerNames()
    outputNames = [layernames[i - 1] for i in net.getUnconnectedOutLayers()]
    outputs = net.forward(outputNames)

    # Prepare to collect results
    hT, wT, _ = img.shape
    bbox, classIds, confs = [], [], []
    count_dict = defaultdict(int)

    # Loop over detections
    for output in outputs:
        for det in output:
            scores = det[5:]  # Confidence for each class
            classId = np.argmax(scores)
            confidence = scores[classId]
            if confidence > confThreshold:
                # Convert detection box to pixel coords
                w, h = int(det[2] * wT), int(det[3] * hT)
                x, y = int(det[0] * wT - w / 2), int(det[1] * hT - h / 2)
                bbox.append([x, y, w, h])
                classIds.append(classId)
                confs.append(float(confidence))

    # Apply non-max suppression to remove duplicates
    indices = cv2.dnn.NMSBoxes(bbox, confs, confThreshold, nmsThreshold)
    for i in indices:
        i = i[0] if isinstance(i, (list, tuple, np.ndarray)) else i
        label = classNames[classIds[i]]
        count_dict[label] += 1
        # Draw bounding box + label on image
        x, y, w, h = bbox[i]
        cv2.rectangle(img, (x, y), (x+w, y+h), (255, 0, 255), 2)
        cv2.putText(img, f"{label}", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 0, 255), 2)
    return img, count_dict

# --- Baseline ---
baseline_counts = None   # Keeps track of last detected objects (for missing detection)

def log_and_publish_async(prev_time, curr_time, missing_objs, current_counts, summary_msg):
    # Run logging + MQTT publishing in a background thread
    try:
        log_to_excel_and_supabase(prev_time, curr_time, missing_objs, current_counts)
        publish_mqtt(summary_msg)
    except Exception as e:
        print("[Async Log Error]:", e)

def send_result(current_counts, result_img):
    # Compare detections with baseline to find missing objects
    global baseline_counts, latest_result
    missing_objs = []
    if baseline_counts is None:
        baseline_counts = current_counts.copy()
    else:
        for obj in baseline_counts:
            if current_counts.get(obj, 0) < baseline_counts[obj]:
                missing_objs.append(f"{obj}:{baseline_counts[obj] - current_counts.get(obj, 0)}")
        baseline_counts = current_counts.copy()

    # Prepare message for ESP32 over serial
    detected_str = " ".join([f"{k}:{v}" for k, v in current_counts.items()]) or "None"
    missing_str = " ".join(missing_objs) or "None"
    msg = f"<Detected:{detected_str}|Missing:{missing_str}>\n"
    ser.write(msg.encode())

    # Update latest result for dashboard
    with lock:
        prev_time = latest_result["curr_time"]
        latest_result["detected_html"] = format_for_dashboard(current_counts)
        latest_result["missing_html"] = format_missing_for_dashboard(missing_objs)
        if latest_result["prev_img"] is not None:
            del latest_result["prev_img"]
        latest_result["prev_img"] = latest_result["curr_img"]
        latest_result["prev_time"] = prev_time
        latest_result["curr_img"] = result_img.copy()
        curr_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        latest_result["curr_time"] = curr_time

    # Summary message for logs + MQTT
    summary_msg = f"{curr_time} | Detected: {detected_str} | Missing: {missing_str}"
    threading.Thread(
        target=log_and_publish_async,
        args=(prev_time, curr_time, missing_objs, current_counts, summary_msg),
        daemon=True
    ).start()

# --- Flask Thread ---
def run_flask():
    # Run Flask server in background
    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)

# Start Flask thread
threading.Thread(target=run_flask, daemon=True).start()

# --- Main Loop ---
last_detection_time = time.time()
while True:
    request_image()              # Ask ESP32 for new image
    img = receive_image()        # Receive image
    if img is None:
        print("[Info] No image received, retrying...")
        time.sleep(1)
        continue

    cv2.imshow("Live YOLO Detection", img)   # Show raw image

    current_time = time.time()
    if current_time - last_detection_time >= 30:   # Run YOLO every 30s
        result_img, counts = detect_and_count(img.copy())   # Detect objects
        send_result(counts, result_img)                     # Send + log results
        cv2.imshow("Live YOLO Detection", result_img)       # Show image with boxes
        last_detection_time = current_time

    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):    # Quit if 'q' pressed
        break
    if key == ord('r'):    # Reset baseline if 'r' pressed
        baseline_counts = None
        print("[Python] Baseline reset.")

# Cleanup when program ends
cv2.destroyAllWindows()
ser.close()
mqtt_client.loop_stop()
mqtt_client.disconnect()

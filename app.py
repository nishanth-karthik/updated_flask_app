# --- Libraries Used ---
import cv2
import numpy as np
import time
from collections import defaultdict
from flask import Flask, render_template, Response, jsonify, request
import threading
from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from supabase import create_client, Client
import paho.mqtt.client as mqtt

# ==========================
# --- Supabase Configuration ---
# ==========================
SUPABASE_URL = "https://osrduqrhujdyvdbciccw.supabase.co/"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9zcmR1cXJodWpkeXZkYmNpY2N3Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc1Njk2NDMyOSwiZXhwIjoyMDcyNTQwMzI5fQ.B3mBxMVL0LW7pCMYdXLZAyau0sR4QpmtuUsjKB5mBxY"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ==========================
# --- YOLO Parameters ---
# ==========================
whT = 224
confThreshold = 0.3
nmsThreshold = 0.3
with open("coco.names", 'rt') as f:
    classNames = f.read().rstrip('\n').split('\n')
net = cv2.dnn.readNetFromDarknet("yolov3-tiny.cfg", "yolov3-tiny.weights")
net.setPreferableBackend(cv2.dnn.DNN_BACKEND_OPENCV)
net.setPreferableTarget(cv2.dnn.DNN_TARGET_CPU)

# ==========================
# --- Flask App Setup ---
# ==========================
app = Flask(__name__)
latest_result = {
    "detected_html": "Waiting...",
    "missing_html": "",
    "prev_img": None,
    "curr_img": None,
    "prev_time": None,
    "curr_time": None
}
lock = threading.Lock()

# ==========================
# --- Excel Logging ---
# ==========================
EXCEL_FILE = "detection_log.xlsx"

def _safe_join_missing(missing_list):
    return ", ".join(missing_list) if missing_list else "No missing objects"

def _safe_join_detected(counts_dict):
    return ", ".join([f"{k}:{v}" for k, v in counts_dict.items()]) if counts_dict else "Nothing detected"

def log_to_excel_and_supabase(prev_time, curr_time, missing_objs, current_counts):
    date_str = curr_time.split(" ")[0] if curr_time else ""
    prev_time_no_date = prev_time.split(" ")[1] if prev_time else ""
    curr_time_no_date = curr_time.split(" ")[1] if curr_time else ""

    missing_str = _safe_join_missing(missing_objs)
    detected_str = _safe_join_detected(current_counts)

    summary = (
        f"At {curr_time_no_date} Missing: {missing_str}\n"
        f"Newly Detected: {detected_str}"
    )

    row = {
        "Date": date_str,
        "Previous Image Timestamp": prev_time_no_date,
        "Current Image Timestamp": curr_time_no_date,
        "Missing Objects": missing_str,
        "Newly Detected": detected_str,
        "Summary": summary
    }

    if os.path.exists(EXCEL_FILE):
        try:
            df = pd.read_excel(EXCEL_FILE)
        except Exception:
            df = pd.DataFrame()
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    else:
        df = pd.DataFrame([row])
    df.to_excel(EXCEL_FILE, index=False)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for cell in ws["F"]:
        cell.alignment = Alignment(wrapText=True)
    wb.save(EXCEL_FILE)

    try:
        supabase.table("detection_log_excel").insert({
            "date": date_str,
            "previous_image_timestamp": prev_time_no_date,
            "current_image_timestamp": curr_time_no_date,
            "missing_objects": missing_str,
            "newly_detected": detected_str,
            "summary": summary
        }).execute()
        print("[Supabase] Row inserted successfully into detection_log_excel")
    except Exception as e:
        print("[Supabase] Insert error:", e)

# ==========================
# --- MQTT Secure (Adafruit IO) ---
# ==========================
BROKER = "io.adafruit.com"
PORT = 8883
USERNAME = "nishanth_11"
PASSWORD = "<YOUR_ADAFRUIT_IO_KEY>"
FEED = "yolo-detection"
TOPIC = f"{USERNAME}/feeds/{FEED}"

mqtt_client = mqtt.Client(client_id="cloud-publisher", protocol=mqtt.MQTTv311)
mqtt_client.username_pw_set(USERNAME, PASSWORD)
mqtt_client.tls_set()
mqtt_client.tls_insecure_set(False)
mqtt_client.connect(BROKER, PORT, keepalive=60)
mqtt_client.loop_start()
print(f"[MQTT] Securely connected to Adafruit IO ({BROKER}), publishing on {TOPIC}")

def publish_mqtt(summary):
    try:
        mqtt_client.publish(TOPIC, summary)
        print(f"[MQTT] Published: {summary}")
    except Exception as e:
        print("[MQTT Error]", e)

# ==========================
# --- YOLO Detection ---
# ==========================
def detect_and_count(img):
    blob = cv2.dnn.blobFromImage(img, 1/255, (whT, whT), [0, 0, 0], 1, crop=False)
    net.setInput(blob)
    layernames = net.getLayerNames()
    outputNames = [layernames[i - 1] for i in net.getUnconnectedOutLayers()]
    outputs = net.forward(outputNames)

    hT, wT, _ = img.shape
    bbox, classIds, confs = [], [], []
    count_dict = defaultdict(int)

    for output in outputs:
        for det in output:
            scores = det[5:]
            classId = np.argmax(scores)
            confidence = scores[classId]
            if confidence > confThreshold:
                w, h = int(det[2] * wT), int(det[3] * hT)
                x, y = int(det[0] * wT - w / 2), int(det[1] * hT - h / 2)
                bbox.append([x, y, w, h])
                classIds.append(classId)
                confs.append(float(confidence))

    indices = cv2.dnn.NMSBoxes(bbox, confs, confThreshold, nmsThreshold)
    for i in indices:
        i = i[0] if isinstance(i, (list, tuple, np.ndarray)) else i
        label = classNames[classIds[i]]
        count_dict[label] += 1
        x, y, w, h = bbox[i]
        cv2.rectangle(img, (x, y), (x+w, y+h), (255, 0, 255), 2)
        cv2.putText(img, f"{label}", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 0, 255), 2)
    return img, count_dict

# ==========================
# --- Baseline Compare ---
# ==========================
baseline_counts = None

def send_result(current_counts, result_img):
    global baseline_counts, latest_result
    missing_objs = []
    if baseline_counts is None:
        baseline_counts = current_counts.copy()
    else:
        for obj in baseline_counts:
            if current_counts.get(obj, 0) < baseline_counts[obj]:
                missing_objs.append(f"{obj}:{baseline_counts[obj] - current_counts.get(obj, 0)}")
        baseline_counts = current_counts.copy()

    detected_str = " ".join([f"{k}:{v}" for k, v in current_counts.items()]) or "None"
    missing_str = " ".join(missing_objs) or "None"

    with lock:
        prev_time = latest_result["curr_time"]
        latest_result["detected_html"] = "<br>".join([f"{k}:{v}" for k,v in current_counts.items()]) or "None"
        latest_result["missing_html"] = "<br>".join(missing_objs) or "None"

        latest_result["prev_img"] = latest_result["curr_img"]
        latest_result["prev_time"] = prev_time
        latest_result["curr_img"] = result_img.copy()
        curr_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        latest_result["curr_time"] = curr_time

    summary_msg = f"{curr_time} | Detected: {detected_str} | Missing: {missing_str}"
    threading.Thread(
        target=lambda: log_to_excel_and_supabase(prev_time, curr_time, missing_objs, current_counts),
        daemon=True
    ).start()
    publish_mqtt(summary_msg)

# ==========================
# --- Flask Routes ---
# ==========================
@app.route('/')
def index():
    return render_template("index4.html")

@app.route('/data')
def data():
    with lock:
        result = {
            "detected_html": latest_result["detected_html"],
            "missing_html": latest_result["missing_html"],
            "prev_img": "/prev.jpg" if latest_result["prev_img"] is not None else "",
            "curr_img": "/curr.jpg" if latest_result["curr_img"] is not None else "",
            "prev_time": latest_result["prev_time"] if latest_result["prev_time"] else "",
            "curr_time": latest_result["curr_time"] if latest_result["curr_time"] else ""
        }
    return jsonify(result)

@app.route('/prev.jpg')
def prev_img():
    with lock:
        if latest_result["prev_img"] is None:
            return "No previous image", 404
        _, buffer = cv2.imencode('.jpg', latest_result["prev_img"])
    return Response(buffer.tobytes(), mimetype='image/jpeg')

@app.route('/curr.jpg')
def curr_img():
    with lock:
        if latest_result["curr_img"] is None:
            return "No current image", 404
        _, buffer = cv2.imencode('.jpg', latest_result["curr_img"])
    return Response(buffer.tobytes(), mimetype='image/jpeg')

# ==========================
# --- Cloud API Endpoint for Upload ---
# ==========================
@app.route('/upload', methods=['POST'])
def upload():
    """
    ESP32 or local client will POST an image here:
    curl -X POST -F "file=@image.jpg" https://<your-cloud-app>/upload
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    npimg = np.frombuffer(file.read(), np.uint8)
    img = cv2.imdecode(npimg, cv2.IMREAD_COLOR)
    if img is None:
        return jsonify({"error": "Invalid image"}), 400

    result_img, counts = detect_and_count(img.copy())
    send_result(counts, result_img)
    return jsonify({"message": "Image processed successfully", "counts": counts})

# ==========================
# --- Start Flask ---
# ==========================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
